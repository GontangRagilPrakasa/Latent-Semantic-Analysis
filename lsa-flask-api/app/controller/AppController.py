from app import app
from openpyxl import load_workbook
from flask import request, jsonify,send_file
from app.constant import RequestMethod
from app.model.QueriesModel import Queries
from app.model.DetailsModel import Details
from app.module.Engine import Engine, preprocess
import os
import requests
import pandas as pd
from math import log
import numpy as np
from io import BytesIO


pd.set_option('display.max_colwidth', -1)

# Load dataset first when open app from browser
dataset = pd.read_excel("app/db/preprocessed-dataset.xlsx")


@app.route("/", methods=RequestMethod.GET)
def index():
    return jsonify({"message": "ok"})

@app.route("/cariquery", methods=RequestMethod.GET)
def cariquery():
    response = list()  # Define response
    if request.method == "POST":
        if "files" in request.files:
            file = request.files["files"]
            file.save(os.path.join("app/db", "queries.xlsx"))
            queries = pd.read_excel("app/db/queries.xlsx")
            queries = queries["Queries"].values
        else:
            resp = {
                "error": "invalid request",
                "path": "/cariquery",
                "message": "request should be file"
            }
            resp = jsonify(resp)
            resp.status_code = 400
            print(resp)
            return resp

    elif request.method == "GET":
        if 'q' in request.args:
            queries = [request.args['q']]
        else:
            resp = {
                "error": "invalid request",
                "path": "/cariquery",
                "message": "request should be query"
            }
            resp = jsonify(resp)
            resp.status_code = 400
            return resp


    # Preproces queries
    queriesPre = list()
    for query in queries:
        queriesPre.append(preprocess(query))

    # Cek di database apakah ada data dengan query pada inputan ataupun file
    for query in queriesPre:
        data = Queries.findByQueryName(query)
        if data is not None:
            response.append(data)

    if len(response) is not 0:
        return jsonify(response)
    else:
        engine = Engine()
        docs = [str(x) for x in dataset['preprocessed_judul']]
        documentsName = list()

        for i, doc in enumerate(docs):
            engine.addDocument(doc)
            documentsName.append("Document_{}".format(i + 1))

        for query in queriesPre:
            engine.setQuery(query)  # Set query pencarian

        titlesScores = engine.process_score()
        ScoreDf = (pd.DataFrame(titlesScores)).T
        ScoreDf.columns = queriesPre
        ScoreDf["Documents"] = documentsName
        ScoreDf["Abstrak"] = dataset["preprocessed_abstrak"].values


        dfListed = list()
        for i in queriesPre:
            labels = list()
            for j in ScoreDf[i]:
                if j > 0.000:
                    labels.append(1)
                else:
                    labels.append(0)
            datadf = pd.DataFrame(ScoreDf[i])
            datadf["Documents"] = ScoreDf["Documents"]
            datadf["Labels"] = labels
            datadf["Judul"] = dataset["Judul"].values
            datadf["Reviewer"] = dataset["Reviewer"].values
            datadf["Abstrak"] = ScoreDf["Abstrak"].values
            dfListed.append(datadf.sort_values(by=[i], ascending=False))

        for i, df in enumerate(dfListed):
            dbQuery = Queries(queriesPre[i])
            for j in range(len(df["Documents"])):
                document = df["Documents"][j]
                label = int(df["Labels"][j])
                score = float(df[queriesPre[i]][j])
                judul = df["Judul"][j]
                reviewer = df["Reviewer"][j]
                abstrak = df["Abstrak"][j]
                data = document, label, score, judul, reviewer, abstrak
                details = Details(data)
                dbQuery.details.append(details)
            dbQuery.save()

        for query in queriesPre:
            data = Queries.findByQueryName(query)
            response.append(data)

        return jsonify(response)

@app.route("/test", methods=RequestMethod.GET)
def getData():
    response = Queries.getAll()
    print(response)
    return jsonify(response)

@app.route("/proses", methods=RequestMethod.GET_POST)
def proses():
    if request.method == "POST":
        if "files" in request.files:
            file = request.files["files"]
            file.save(os.path.join("app/db", "datasetbaru.xlsx"))
        else:
            resp = {
                "error": "invalid request",
                "path": "/proses",
                "message": "request should be file"
            }
      
    databr = pd.read_excel("app/db/datasetbaru.xlsx")
    preprocessed_jd = []
    judul = databr.Judul.tolist()
    preprocessed_abst = []
    abst = databr.Abstrak.tolist()

    for doc in judul:
        preprocessed_jd.append(preprocess(doc))
    databr['preprocessed_judul'] = preprocessed_jd
    databr.to_excel("app/db/datasetbaru.xlsx", engine="xlsxwriter", index=False)
    
    for docs in abst:
        preprocessed_abst.append(preprocess(docs))
    databr['preprocessed_abstrak'] = preprocessed_abst 
    databr.to_excel("app/db/datasetbaru.xlsx", engine="xlsxwriter", index=False)

        #datalama-panggil
    wb = load_workbook(filename = 'app/db/preprocessed-dataset.xlsx')
    sheet_ranges = wb['Sheet1']
    dat = pd.DataFrame(sheet_ranges.values)
    dat.columns = ['Judul','Reviewer','Abstrak','preprocessed_judul','preprocessed_abstrak']
    data = dat[1:15000]

    asli = []
    for dt in data['Judul']:
        asli.append(dt)
    asli1 = []
    for dt1 in data['Reviewer']:
        asli1.append(dt1)
    asli2 = []
    for dt2 in data['Abstrak']:
        asli2.append(dt2)
    asli3 = []
    for dt3 in data['preprocessed_judul']:
        asli3.append(dt3)
    asli4 = []
    for dt4 in data['preprocessed_abstrak']:
        asli4.append(dt4)


    workb = load_workbook(filename='app/db/datasetbaru.xlsx')
    sheet_ranges1 = workb['Sheet1']
    dat2 = pd.DataFrame(sheet_ranges1.values)
    dat2.columns = ['Judul','Reviewer','Abstrak','preprocessed_judul','preprocessed_abstrak']
    data2 = dat2[1:500]
        
    rows = []
    for d in data2['Judul']:
        rows.append(d)
    rows1 = []
    for d1 in data2['Reviewer']:
        rows1.append(d1)
    rows2 = []
    for d2 in data2['Abstrak']:
        rows2.append(d2)
    rows3 = []
    for d3 in data2['preprocessed_judul']:
        rows3.append(d3)
    rows4 = []
    for d4 in data2['preprocessed_abstrak']:
        rows4.append(d4)


    for row in rows:
        asli.append(row)
    for row1 in rows1:
        asli1.append(row1)
    for row2 in rows2:
        asli2.append(row2)
    for row3 in rows3:
        asli3.append(row3)
    for row4 in rows4:
        asli4.append(row4)

    datasetku = pd.read_excel("app/db/a.xlsx")
    datasetku['Judul'] = asli
    datasetku['Reviewer'] = asli1
    datasetku['Abstrak'] = asli2
    datasetku['preprocessed_judul'] = asli3
    datasetku['preprocessed_abstrak'] = asli4

    datasetku = datasetku.to_excel("app/db/preprocessed-dataset.xlsx", engine="xlsxwriter", index=False)
    return jsonify({"message": "Data sudah ditambahkan"})